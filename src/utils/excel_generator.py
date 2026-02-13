from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from src.data.models.batch import Batch
from src.data.models.product import Product


def generate_batch_report_excel(
    batch: Batch,
    products: Iterable[Product],
    file_path: str,
) -> None:
    """
    Формирует Excel-отчёт по партии с тремя листами:
    1. Информация о партии
    2. Продукция
    3. Статистика (упрощённая)
    """
    from openpyxl import Workbook

    wb = Workbook()

    # Лист 1: Информация о партии
    ws_info = wb.active
    ws_info.title = "Информация о партии"
    ws_info.append(["Номер партии", batch.batch_number])
    ws_info.append(["Дата партии", batch.batch_date.isoformat()])
    ws_info.append(["Статус", "Закрыта" if batch.is_closed else "Открыта"])
    ws_info.append(["Рабочий центр", batch.work_center.name if batch.work_center else ""])
    ws_info.append(["Смена", batch.shift])
    ws_info.append(["Бригада", batch.team])
    ws_info.append(["Номенклатура", batch.nomenclature])
    ws_info.append(["Начало смены", batch.shift_start])
    ws_info.append(["Окончание смены", batch.shift_end])

    # Лист 2: Продукция
    ws_prod = wb.create_sheet(title="Продукция")
    ws_prod.append(["ID", "Уникальный код", "Аггрегирована", "Дата аггрегации"])
    for p in products:
        ws_prod.append(
            [
                p.id,
                p.unique_code,
                "Да" if p.is_aggregated else "Нет",
                p.aggregated_at or "-",
            ]
        )

    # Лист 3: Статистика (упрощённо считаем по коллекции)
    products_list = list(products)
    total = len(products_list)
    aggregated = sum(1 for p in products_list if p.is_aggregated)
    remaining = total - aggregated
    rate = round(aggregated / total * 100, 2) if total else 0.0

    ws_stats = wb.create_sheet(title="Статистика")
    ws_stats.append(["Всего продукции", total])
    ws_stats.append(["Аггрегировано", aggregated])
    ws_stats.append(["Осталось", remaining])
    ws_stats.append(["Процент выполнения", rate])

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)


def generate_batch_report_pdf(
    batch: Batch,
    products: Iterable[Product],
    file_path: str,
) -> None:
    """
    Упрощённый PDF-отчёт с основной информацией и краткой статистикой.
    """
    products_list = list(products)
    total = len(products_list)
    aggregated = sum(1 for p in products_list if p.is_aggregated)
    remaining = total - aggregated
    rate = round(aggregated / total * 100, 2) if total else 0.0

    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4

    y = height - 50
    lines = [
        f"Отчёт по партии #{batch.batch_number} от {batch.batch_date.isoformat()}",
        "",
        f"Рабочий центр: {batch.work_center.name if batch.work_center else ''}",
        f"Смена: {batch.shift}",
        f"Бригада: {batch.team}",
        f"Номенклатура: {batch.nomenclature}",
        "",
        f"Всего продукции: {total}",
        f"Аггрегировано: {aggregated}",
        f"Осталось: {remaining}",
        f"Процент выполнения: {rate}%",
        "",
        f"Сгенерировано: {datetime.now()}",
    ]

    for line in lines:
        c.drawString(40, y, line)
        y -= 20

    c.showPage()
    c.save()


def export_batches_to_excel(rows: list[dict], file_path: str) -> None:
    df = pd.DataFrame(rows)
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(file_path, index=False)


def export_batches_to_csv(rows: list[dict], file_path: str) -> None:
    df = pd.DataFrame(rows)
    Path(file_path).parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(file_path, index=False, encoding="utf-8-sig")


